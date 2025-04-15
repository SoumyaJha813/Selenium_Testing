import time

import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()

driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
#download the excel file from webpage
#make changes in the file in local
#upload to browser again from local
#wait for the message toast to appear
#wait for msg to dissapper
#assert that value is properly updated

#fetch the text apple price from webpage
fruit_name = "Apple"
filepath = "C:\\Users\\sojha\\Downloads\\download.xlsx"
time.sleep(3)
#colName - price
new_value = "990"
#file downloaded
driver.find_element(By.ID, "downloadButton").click()


def update_excel_data(filepath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active

    Dict = {}
    #get the apple price from excel sheet
    for i in range(1, sheet.max_row + 1):
        for x in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=x).value == colName:
                Dict["col"] = x
            if sheet.cell(row=i, column=x).value == searchTerm:
                Dict["row"] = i
    print("apple price value from excel:", sheet.cell(row=Dict["row"], column=Dict["col"]).value)
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(filepath)


#call the update excel function
update_excel_data(filepath, fruit_name, "price", new_value)

#file upload
file_input = driver.find_element(By.XPATH, "//input[@type='file']")
file_input.send_keys(filepath)

#wait for msg to appear
wait = WebDriverWait(driver, 5)
toast_locator = By.XPATH, "//div[@class='Toastify__toast-body']/div[2]"
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
price_fruit = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text

assert price_fruit == new_value

time.sleep(5)
