import openpyxl
book = openpyxl.load_workbook("C:\\Users\\sojha\\Documents\\PythonDemo.xlsx")

sheet = book.active
#getting the cell value from excel sheet
cell = sheet.cell(row=1, column=2)
print(cell.value)

#putting the value in excel sheet
sheet.cell(row=2, column=2).value = "Soumya"
print(sheet.cell(row=2, column=2).value)

#get max number of rows
print("Max row:", sheet.max_row)
#get max number of columns
print("Max column:", sheet.max_column)
print("Printing rows and column: \n")
print(sheet['A5'].value)
Dict = {}
list_names = []
for i in range(2, sheet.max_row+1):
    #if sheet.cell(row=i, column=1).value == "Testcase2":
    Dict = {}
    for j in range(2, sheet.max_column+1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        list_names.append({sheet.cell(row=1, column=j).value : sheet.cell(row=i, column=j).value})
        #list_names.append(Dict)

print(Dict)
print(list_names)